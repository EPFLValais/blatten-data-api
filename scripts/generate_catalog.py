#!/usr/bin/env python3
"""
Blatten STAC Catalog Generator

Generates STAC 1.1.0 catalog with hierarchical items:
- Dataset items (from Excel) with combined geometry and archive asset
- File items (individual files) with individual geometry, linked to parent

Usage:
    uv run generate_catalog.py
    uv run generate_catalog.py --xlsx metadata.xlsx --data /path/to/FINAL_Data
"""

import argparse
import json
import os
import re
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

# =============================================================================
# Progress Bar (tqdm or fallback)
# =============================================================================

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

class SimpleProgress:
    """Simple progress bar fallback when tqdm is not available."""
    def __init__(self, iterable=None, total=None, desc="", unit="it", leave=True):
        self.iterable = iterable
        self.total = total or (len(iterable) if iterable and hasattr(iterable, '__len__') else None)
        self.desc = desc
        self.unit = unit
        self.n = 0
        self.leave = leave
        self._last_print_len = 0

    def __iter__(self):
        if self.iterable is None:
            return
        for item in self.iterable:
            yield item
            self.update(1)
        if self.leave:
            print()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        if self.leave:
            print()

    def update(self, n=1):
        self.n += n
        self._print_progress()

    def _print_progress(self):
        if self.total:
            pct = (self.n / self.total) * 100
            bar_len = 30
            filled = int(bar_len * self.n / self.total)
            bar = "█" * filled + "░" * (bar_len - filled)
            line = f"\r  {self.desc}: {bar} {self.n}/{self.total} ({pct:.0f}%)"
        else:
            line = f"\r  {self.desc}: {self.n} {self.unit}"
        # Clear previous line if shorter
        print(line + " " * max(0, self._last_print_len - len(line)), end="", flush=True)
        self._last_print_len = len(line)

    def set_postfix_str(self, s):
        pass  # Ignored in simple mode

def progress_bar(iterable=None, total=None, desc="", unit="it", leave=True):
    """Create a progress bar, using tqdm if available."""
    if HAS_TQDM:
        return tqdm(iterable, total=total, desc=f"  {desc}", unit=unit, leave=leave,
                    bar_format="{desc}: {bar:30} {n_fmt}/{total_fmt} ({percentage:.0f}%)")
    return SimpleProgress(iterable, total, desc, unit, leave)

# =============================================================================
# Configuration
# =============================================================================

STAC_VERSION = "1.1.0"
CATALOG_ID = "birch-glacier-collapse"
CATALOG_TITLE = "Birch Glacier Collapse Dataset"
CATALOG_DESCRIPTION = "Multi-sensor monitoring data from the May 28, 2025 Birch glacier collapse at Blatten, Switzerland."
LICENSE = "CC-BY-NC-SA-4.0"
DEFAULT_BBOX = [7.75, 46.38, 7.90, 46.45]
POINT_BUFFER = 0.001
MAX_WORKERS = os.cpu_count() or 4  # Parallel geometry extraction workers

STAC_BASE_URL = "${STAC_BASE_URL}"
S3_BASE_URL = "${S3_BASE_URL}"

PROVIDERS = [
    {"name": "Canton du Valais", "roles": ["licensor"], "url": "https://www.vs.ch/"},
    {"name": "EPFL ALPOLE", "roles": ["processor", "host"], "url": "https://www.epfl.ch/research/domains/alpole/"},
]

PRODUCT_TO_COLLECTION = {
    "A": "webcam-image", "B": "deformation-analysis", "D": "orthophoto",
    "E": "radar-displacement", "F": "radar-amplitude", "G": "radar-coherence",
    "H": "radar-velocity", "I": "dsm", "J": "dem", "K": "point-cloud",
    "L": "3d-model", "M": "gnss-data", "N": "thermal-image",
    "O": "hydrology", "P": "lake-level", "U": "radar-timeseries",
}

MIME_TYPES = {
    ".tif": "image/tiff; application=geotiff",
    ".tiff": "image/tiff; application=geotiff",
    ".laz": "application/vnd.laszip",
    ".las": "application/vnd.las",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".zip": "application/zip",
    ".csv": "text/csv",
    ".json": "application/json",
    ".obj": "model/obj",
    ".ply": "application/ply",
}

DNAGE_SHARED = {
    "10M_11M_GNSS": ["10Ma00", "11Ma00", "11Mb00", "11Mc00", "11Md00", "11Me00", "11Mf00", "11Mg00", "11Mh00"],
    "17_LakeLevel_Geoazimut": ["17Pa00", "17Pb00"],
    "18_LonzaRiverDischarge_upstream": ["18Oa01", "18Oa02", "18Ob01", "18Ob02", "18Ob03"],
    "19_LonzaRiverDischarge_Blatten": ["19Oa00", "19Ob00"],
    "20_LonzaRiverDischarge_Kippel": ["20Oa00", "20Ob01", "20Ob02", "20Ob03"],
    "21_LonzaRiverDischarge_Gampel": ["21Oa00", "21Ob00", "21Oc00"],
}

SCRIPT_DIR = Path(__file__).parent.resolve()
ROOT = SCRIPT_DIR.parent
OUTPUT_PATH = ROOT / "stac"

# =============================================================================
# Data Model
# =============================================================================

@dataclass
class FileInfo:
    """Individual file with its geometry and size."""
    path: Path
    size: int = 0
    bbox: list = None
    geometry: dict = None

@dataclass
class Dataset:
    """Dataset from Excel with its files."""
    code: str
    sensor: str = ""
    dataset: str = ""
    description: str = ""
    source: str = ""
    date_first: str = ""
    date_last: str = ""
    collection_id: str = ""
    folder_path: Path = None
    files: list = field(default_factory=list)  # List of FileInfo
    bbox: list = None  # Combined bbox
    geometry: dict = None  # Combined geometry
    archive_size: int = 0  # Actual .zip file size (0 if not found)

# =============================================================================
# Coordinate Transform
# =============================================================================

_transformer = None

def lv95_to_wgs84(x: float, y: float):
    global _transformer
    if _transformer is None:
        try:
            from pyproj import Transformer
            _transformer = Transformer.from_crs("EPSG:2056", "EPSG:4326", always_xy=True)
        except ImportError:
            return None
    return _transformer.transform(x, y)

# =============================================================================
# File Discovery
# =============================================================================

def find_files(pattern: str, search_path: Path) -> list:
    return sorted(search_path.rglob(pattern))

def select_file(files: list, prompt: str) -> Path:
    if not files:
        return None
    if len(files) == 1:
        print(f"{prompt}: {files[0]}")
        return files[0]
    print(f"\n{prompt}:")
    for i, f in enumerate(files, 1):
        print(f"  {i}. {f}")
    while True:
        try:
            choice = input("Select [1]: ").strip() or "1"
            idx = int(choice) - 1
            if 0 <= idx < len(files):
                return files[idx]
        except (ValueError, KeyboardInterrupt):
            pass
        print("Invalid selection")

# =============================================================================
# Geometry Extraction
# =============================================================================

def extract_tiff_geometry(path: Path) -> dict:
    """Extract geometry from GeoTIFF."""
    try:
        r = subprocess.run(["gdalinfo", "-json", str(path)], capture_output=True, text=True, timeout=30)
        if r.returncode == 0:
            info = json.loads(r.stdout)
            if "wgs84Extent" in info:
                coords = info["wgs84Extent"]["coordinates"][0]
                lons, lats = [c[0] for c in coords], [c[1] for c in coords]
                return {"bbox": [min(lons), min(lats), max(lons), max(lats)], "geometry": info["wgs84Extent"]}
    except Exception:
        pass
    return None

def extract_laz_geometry(path: Path) -> dict:
    """Extract geometry from LAZ/LAS point cloud."""
    try:
        r = subprocess.run(["pdal", "info", "--summary", str(path)], capture_output=True, text=True, timeout=60)
        if r.returncode == 0:
            info = json.loads(r.stdout)
            bounds = info.get("summary", {}).get("bounds", {})
            if bounds:
                minx, maxx = bounds.get("minx"), bounds.get("maxx")
                miny, maxy = bounds.get("miny"), bounds.get("maxy")
                if minx and maxx and miny and maxy:
                    # Check if LV95 (Swiss coordinates)
                    if minx > 2000000 and miny > 1000000:
                        sw = lv95_to_wgs84(minx, miny)
                        ne = lv95_to_wgs84(maxx, maxy)
                        if sw and ne:
                            return {
                                "bbox": [sw[0], sw[1], ne[0], ne[1]],
                                "geometry": {
                                    "type": "Polygon",
                                    "coordinates": [[[sw[0], sw[1]], [ne[0], sw[1]], [ne[0], ne[1]], [sw[0], ne[1]], [sw[0], sw[1]]]]
                                }
                            }
                    else:
                        return {
                            "bbox": [minx, miny, maxx, maxy],
                            "geometry": {
                                "type": "Polygon",
                                "coordinates": [[[minx, miny], [maxx, miny], [maxx, maxy], [minx, maxy], [minx, miny]]]
                            }
                        }
    except Exception:
        pass
    return None

def extract_file_geometry(path: Path) -> dict:
    """Extract geometry from a file based on its type."""
    suffix = path.suffix.lower()
    if suffix in (".tif", ".tiff"):
        return extract_tiff_geometry(path)
    elif suffix in (".laz", ".las"):
        return extract_laz_geometry(path)
    return None

def combine_bboxes(bboxes: list) -> list:
    """Combine multiple bboxes into one encompassing bbox."""
    if not bboxes:
        return None
    return [
        min(b[0] for b in bboxes),
        min(b[1] for b in bboxes),
        max(b[2] for b in bboxes),
        max(b[3] for b in bboxes)
    ]

def bbox_to_polygon(bbox: list) -> dict:
    """Convert bbox to GeoJSON Polygon."""
    return {
        "type": "Polygon",
        "coordinates": [[
            [bbox[0], bbox[1]], [bbox[2], bbox[1]],
            [bbox[2], bbox[3]], [bbox[0], bbox[3]],
            [bbox[0], bbox[1]]
        ]]
    }

# =============================================================================
# Excel Parsing
# =============================================================================

def parse_excel(path: Path) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = next((n for n in ["Data", "Test_Data", "All_Data"] if n in wb.sheetnames), wb.sheetnames[0])
    print(f"(using sheet: {sheet})")
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()

    hdr_idx = next((i for i, r in enumerate(rows) if r and str(r[0]).strip().lower() == "code"), None)
    if hdr_idx is None:
        print(f"ERROR: No 'Code' header found. First rows:")
        for i, r in enumerate(rows[:5]):
            if r:
                print(f"  {i}: {r[0]}")
        return []

    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[hdr_idx]]

    def get(row, *keys):
        for k in keys:
            if k in headers:
                v = row[headers.index(k)]
                if v is not None:
                    return v.strftime("%Y-%m-%d") if isinstance(v, datetime) else str(v).strip()
        return ""

    datasets = []
    for row in rows[hdr_idx + 1:]:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        if not code or len(code) < 3:
            continue
        datasets.append(Dataset(
            code=code,
            sensor=get(row, "sensor"),
            dataset=get(row, "dataset"),
            description=get(row, "description"),
            source=get(row, "source"),
            date_first=get(row, "date_first_(provided)"),
            date_last=get(row, "date_last_(provided)"),
            collection_id=PRODUCT_TO_COLLECTION.get(code[2].upper(), ""),
        ))
    return datasets

# =============================================================================
# Folder Scanning
# =============================================================================

def scan_dataset_folders(data_path: Path, datasets: list) -> dict:
    """Map dataset codes to their folder paths."""
    code_to_folder = {}

    for provider in ["DNAGE", "Geopraevent", "Terradata"]:
        pdir = data_path / provider
        if not pdir.exists():
            continue
        for entry in pdir.iterdir():
            if entry.is_dir():
                code = entry.name.split("_")[0]
                code_to_folder[code] = entry
                # Nested subdirs
                for sub in entry.iterdir():
                    if sub.is_dir() and re.match(r"^\d{2}[A-Z][a-z]\d{2}$", sub.name.split("_")[0]):
                        code_to_folder[sub.name.split("_")[0]] = sub
            elif entry.suffix.lower() in (".tif", ".tiff"):
                code_to_folder[entry.stem] = entry

    # DNAGE shared folders
    for folder, codes in DNAGE_SHARED.items():
        fp = data_path / "DNAGE" / folder
        if fp.exists():
            for code in codes:
                code_to_folder.setdefault(code, fp)

    return code_to_folder

def scan_files_in_folder(folder: Path) -> list:
    """Get all relevant files in a folder."""
    if folder.is_file():
        return [folder]

    files = []
    extensions = (".tif", ".tiff", ".laz", ".las", ".jpg", ".jpeg", ".png", ".csv", ".json", ".obj", ".ply")
    for ext in extensions:
        files.extend(folder.glob(f"*{ext}"))
        files.extend(folder.glob(f"*{ext.upper()}"))
    return sorted(files)

# =============================================================================
# Sensor Geometry Lookup
# =============================================================================

def get_sensor_geometry(code: str, sensors: dict) -> dict:
    """Lookup geometry from sensor locations file."""
    for sid, s in sensors.get("sensors", {}).items():
        if code in s.get("items", []):
            if s.get("x") and s.get("y"):
                result = lv95_to_wgs84(s["x"], s["y"])
                if result:
                    lon, lat = result
                    return {
                        "bbox": [lon - POINT_BUFFER, lat - POINT_BUFFER, lon + POINT_BUFFER, lat + POINT_BUFFER],
                        "geometry": {"type": "Point", "coordinates": [lon, lat]}
                    }
            if s.get("lon") and s.get("lat"):
                lon, lat = s["lon"], s["lat"]
                return {
                    "bbox": [lon - POINT_BUFFER, lat - POINT_BUFFER, lon + POINT_BUFFER, lat + POINT_BUFFER],
                    "geometry": {"type": "Point", "coordinates": [lon, lat]}
                }
    return None

# =============================================================================
# STAC Generation
# =============================================================================

def build_file_item(dataset: Dataset, file_info: FileInfo, file_idx: int, data_path: Path) -> dict:
    """Build a STAC item for an individual file."""
    file_id = f"{dataset.code}-{file_idx:04d}"
    rel_path = file_info.path.relative_to(data_path) if file_info.path.is_relative_to(data_path) else file_info.path.name

    # Datetime
    if dataset.date_first and dataset.date_last:
        dt, sdt, edt = None, f"{dataset.date_first}T00:00:00Z", f"{dataset.date_last}T23:59:59Z"
    elif dataset.date_first:
        dt, sdt, edt = f"{dataset.date_first}T00:00:00Z", None, None
    else:
        dt, sdt, edt = None, None, None

    props = {
        "title": file_info.path.name,
        "datetime": dt,
        "blatten:dataset": dataset.code,
        "blatten:archive": f"{dataset.code}.zip",
    }
    if dataset.description:
        props["description"] = dataset.description
    if sdt:
        props["start_datetime"] = sdt
        props["end_datetime"] = edt
    if dataset.source:
        props["blatten:source"] = dataset.source

    # Asset
    suffix = file_info.path.suffix.lower()
    mime_type = MIME_TYPES.get(suffix, "application/octet-stream")

    # Build data asset with optional file size
    data_asset = {
        "href": f"{S3_BASE_URL}/{rel_path}",
        "type": mime_type,
        "title": file_info.path.name,
    }
    if file_info.size > 0:
        data_asset["file:size"] = file_info.size

    return {
        "type": "Feature",
        "stac_version": STAC_VERSION,
        "id": file_id,
        "geometry": file_info.geometry,
        "bbox": file_info.bbox,
        "properties": props,
        "links": [
            {"rel": "collection", "href": f"{STAC_BASE_URL}/stac/collections/{dataset.collection_id}"},
            {"rel": "parent", "href": f"{STAC_BASE_URL}/stac/collections/{dataset.collection_id}/items/{dataset.code}"},
        ],
        "assets": {
            "data": data_asset
        },
        "collection": dataset.collection_id,
    }

def build_dataset_item(dataset: Dataset, file_items: list) -> dict:
    """Build a STAC item for a dataset (parent item)."""
    # Datetime
    if dataset.date_first and dataset.date_last:
        dt, sdt, edt = None, f"{dataset.date_first}T00:00:00Z", f"{dataset.date_last}T23:59:59Z"
    elif dataset.date_first:
        dt, sdt, edt = f"{dataset.date_first}T00:00:00Z", None, None
    else:
        dt, sdt, edt = None, None, None

    props = {
        "title": f"{dataset.sensor} - {dataset.dataset}".strip(" -") or dataset.code,
        "datetime": dt,
        "blatten:code": dataset.code,
        "blatten:file_count": len(file_items),
    }
    if dataset.description:
        props["description"] = dataset.description
    if sdt:
        props["start_datetime"] = sdt
        props["end_datetime"] = edt
    if dataset.source:
        props["blatten:source"] = dataset.source

    # Links to children
    links = [
        {"rel": "collection", "href": f"{STAC_BASE_URL}/stac/collections/{dataset.collection_id}"},
    ]
    for fi in file_items:
        links.append({"rel": "item", "href": f"{STAC_BASE_URL}/stac/collections/{dataset.collection_id}/items/{fi['id']}"})

    # Assets - archive (only include file:size if actual .zip exists)
    archive_asset = {
        "href": f"{S3_BASE_URL}/archives/{dataset.code}.zip",
        "type": "application/zip",
        "title": f"{dataset.code} complete archive",
    }
    if dataset.archive_size > 0:
        archive_asset["file:size"] = dataset.archive_size

    assets = {"archive": archive_asset}

    return {
        "type": "Feature",
        "stac_version": STAC_VERSION,
        "id": dataset.code,
        "geometry": dataset.geometry,
        "bbox": dataset.bbox,
        "properties": props,
        "links": links,
        "assets": assets,
        "collection": dataset.collection_id,
    }

def generate_stac(datasets: list, output: Path, show_progress: bool = False) -> tuple:
    """Generate complete STAC catalog."""
    output.mkdir(parents=True, exist_ok=True)
    (output / "collections").mkdir(exist_ok=True)

    # Group by collection
    by_coll = {}
    for ds in datasets:
        if ds.collection_id:
            by_coll.setdefault(ds.collection_id, []).append(ds)

    collections = []
    all_dataset_items = []
    all_file_items = []

    coll_iter = sorted(by_coll.items())
    if show_progress:
        coll_iter = progress_bar(coll_iter, desc="Building collections")

    for cid, coll_datasets in coll_iter:
        coll_dataset_items = []
        coll_file_items = []

        for ds in coll_datasets:
            # File items for this dataset - ALL files, not just those with geometry
            file_items = [build_file_item(ds, fi, idx, output.parent)
                         for idx, fi in enumerate(ds.files, 1)]
            coll_file_items.extend(file_items)

            # Dataset item
            dataset_item = build_dataset_item(ds, file_items)
            coll_dataset_items.append(dataset_item)

        all_dataset_items.extend(coll_dataset_items)
        all_file_items.extend(coll_file_items)

        # Collection extent
        all_items = coll_dataset_items + coll_file_items
        bboxes = [i["bbox"] for i in all_items if i.get("bbox")]
        if bboxes:
            bbox = [combine_bboxes(bboxes)]
        else:
            bbox = [DEFAULT_BBOX]

        dates = sorted([d for ds in coll_datasets for d in (ds.date_first, ds.date_last) if d])
        interval = [[f"{dates[0]}T00:00:00Z", f"{dates[-1]}T23:59:59Z"]] if dates else [[None, None]]
        desc = next((ds.description for ds in coll_datasets if ds.description), f"{cid.replace('-', ' ').title()} data")

        coll = {
            "type": "Collection",
            "id": cid,
            "stac_version": STAC_VERSION,
            "title": cid.replace("-", " ").title(),
            "description": desc,
            "license": LICENSE,
            "providers": PROVIDERS,
            "extent": {"spatial": {"bbox": bbox}, "temporal": {"interval": interval}},
            "links": [{"rel": "self", "href": f"{STAC_BASE_URL}/stac/collections/{cid}"}],
        }
        collections.append(coll)

        # Write collection files
        with open(output / "collections" / f"{cid}.json", "w") as f:
            json.dump(coll, f, indent=2)

        # All items for this collection (datasets + files)
        all_coll_items = coll_dataset_items + coll_file_items
        with open(output / "collections" / f"{cid}_items.json", "w") as f:
            json.dump({"type": "FeatureCollection", "features": all_coll_items}, f, indent=2)

    # Root catalog
    catalog = {
        "type": "Catalog",
        "id": CATALOG_ID,
        "stac_version": STAC_VERSION,
        "title": CATALOG_TITLE,
        "description": CATALOG_DESCRIPTION,
        "links": [{"rel": "child", "href": f"{STAC_BASE_URL}/stac/collections/{c['id']}"} for c in collections]
    }

    with open(output / "catalog.json", "w") as f:
        json.dump(catalog, f, indent=2)

    # All items combined
    all_items = all_dataset_items + all_file_items
    with open(output / "all_items.json", "w") as f:
        json.dump({"type": "FeatureCollection", "features": all_items}, f, indent=2)

    with open(output / "collections.json", "w") as f:
        json.dump({"collections": collections}, f, indent=2)

    return len(collections), len(all_dataset_items), len(all_file_items)

# =============================================================================
# Main
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="Generate STAC catalog from Excel metadata")
    parser.add_argument("--xlsx", type=Path, help="Path to Excel metadata file (.xlsx)")
    parser.add_argument("--data", type=Path, help="Path to data folder containing DNAGE/Geopraevent/Terradata")
    parser.add_argument("--sensors", type=Path, help="Path to sensor_locations.json")
    parser.add_argument("--output", type=Path, help="Output directory for STAC catalog files")
    parser.add_argument("--workers", type=int, default=MAX_WORKERS, help=f"Parallel workers for geometry extraction (default: {MAX_WORKERS})")
    parser.add_argument("--validate-s3", type=str, metavar="URL", help="Validate files exist at S3 proxy URL (e.g., http://localhost:8888/s3)")
    args = parser.parse_args()

    print("Blatten STAC Catalog Generator")
    print(f"(Using {args.workers} parallel workers for geometry extraction)\n")

    # Find and select Excel file
    if args.xlsx:
        xlsx_path = args.xlsx.resolve()
        if not xlsx_path.exists():
            print(f"ERROR: Excel file not found: {xlsx_path}")
            sys.exit(1)
        print(f"Excel metadata file: {xlsx_path}")
    else:
        xlsx_files = find_files("*.xlsx", ROOT)
        xlsx_path = select_file(xlsx_files, "Excel metadata file")
        if not xlsx_path:
            print(f"ERROR: No Excel file found in {ROOT}")
            print(f"Provide path with: --xlsx /path/to/metadata.xlsx")
            sys.exit(1)

    # Find data folder
    data_path = (args.data.resolve() if args.data else ROOT / "data")
    if not data_path.exists():
        print(f"ERROR: Data folder not found: {data_path}")
        print(f"Provide path with: --data /path/to/data")
        sys.exit(1)
    print(f"Data folder: {data_path}")

    # Output path
    output_path = (args.output.resolve() if args.output else OUTPUT_PATH)

    # Load sensor locations
    sensors = {}
    if args.sensors:
        if args.sensors.exists():
            print(f"Sensor locations: {args.sensors}")
            with open(args.sensors) as f:
                sensors = json.load(f)
    else:
        sensor_files = find_files("sensor_locations.json", ROOT)
        if sensor_files:
            print(f"Sensor locations: {sensor_files[0]}")
            with open(sensor_files[0]) as f:
                sensors = json.load(f)

    # Parse Excel
    print(f"\nParsing Excel... ", end="", flush=True)
    datasets = parse_excel(xlsx_path)
    print(f"{len(datasets)} datasets")

    # Map datasets to folders
    print(f"Mapping folders... ", end="", flush=True)
    code_to_folder = scan_dataset_folders(data_path, datasets)
    matched = 0
    for ds in datasets:
        if ds.code in code_to_folder:
            ds.folder_path = code_to_folder[ds.code]
            matched += 1
    print(f"{matched}/{len(datasets)} matched")

    # Scan files
    print(f"\nScanning files...")
    all_files = []  # [(dataset_index, file_path), ...]
    for i, ds in enumerate(progress_bar(datasets, desc="Scanning datasets")):
        if ds.folder_path:
            files = scan_files_in_folder(ds.folder_path)
            for f in files:
                all_files.append((i, f))

    print(f"\n  Total files found: {len(all_files)}")

    # Identify files needing geometry extraction (GeoTIFF and LAZ/LAS)
    geo_files = [(i, f) for i, f in all_files if f.suffix.lower() in (".tif", ".tiff", ".laz", ".las")]
    other_files = [(i, f) for i, f in all_files if f.suffix.lower() not in (".tif", ".tiff", ".laz", ".las")]
    print(f"  Files needing geometry extraction: {len(geo_files)}")
    print(f"  Other files (no geometry): {len(other_files)}")

    # Extract geometry in parallel
    print(f"\nExtracting geometry ({args.workers} workers)...")
    file_geometries = {}  # path -> (bbox, geometry)
    geo_tiff = 0
    geo_laz = 0

    def extract_geometry_task(item):
        """Worker task for parallel geometry extraction."""
        idx, path = item
        geom = extract_file_geometry(path)
        return (idx, path, geom)

    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {executor.submit(extract_geometry_task, item): item for item in geo_files}
        with progress_bar(total=len(futures), desc="Extracting geometry") as pbar:
            for future in as_completed(futures):
                idx, path, geom = future.result()
                if geom:
                    file_geometries[path] = geom
                    if path.suffix.lower() in (".tif", ".tiff"):
                        geo_tiff += 1
                    else:
                        geo_laz += 1
                pbar.update(1)

    print(f"\n  GeoTIFF geometry extracted: {geo_tiff}")
    print(f"  LAZ/LAS geometry extracted: {geo_laz}")

    # Build file lists for each dataset
    print(f"\nBuilding dataset file lists...")
    geo_sensor = 0
    datasets_with_geo = 0

    for i, ds in enumerate(progress_bar(datasets, desc="Processing datasets")):
        file_bboxes = []
        dataset_files = [(idx, f) for idx, f in all_files if idx == i]

        for _, f in dataset_files:
            # Get file size
            try:
                file_size = f.stat().st_size if f.exists() else 0
            except OSError:
                file_size = 0

            fi = FileInfo(path=f, size=file_size)
            if f in file_geometries:
                geom = file_geometries[f]
                fi.bbox = geom["bbox"]
                fi.geometry = geom["geometry"]
                file_bboxes.append(fi.bbox)
            ds.files.append(fi)

        # Combine file bboxes for dataset
        if file_bboxes:
            ds.bbox = combine_bboxes(file_bboxes)
            ds.geometry = bbox_to_polygon(ds.bbox)
            datasets_with_geo += 1
        else:
            # Try sensor locations for datasets without files
            sensor_geom = get_sensor_geometry(ds.code, sensors)
            if sensor_geom:
                ds.bbox = sensor_geom["bbox"]
                ds.geometry = sensor_geom["geometry"]
                datasets_with_geo += 1
                geo_sensor += 1

    print(f"\n  Datasets with geometry from files: {datasets_with_geo - geo_sensor}")
    print(f"  Datasets with geometry from sensors: {geo_sensor}")
    print(f"  Total datasets with geometry: {datasets_with_geo}/{len(datasets)}")

    # Check for archive files and get their sizes
    print(f"\nChecking for archive files...")
    archives_path = data_path / "archives"
    if not archives_path.exists():
        for alt in [data_path.parent / "archives", output_path.parent / "archives"]:
            if alt.exists():
                archives_path = alt
                break

    archives_found = 0
    if archives_path.exists():
        for ds in datasets:
            zip_path = archives_path / f"{ds.code}.zip"
            if zip_path.exists():
                try:
                    ds.archive_size = zip_path.stat().st_size
                    archives_found += 1
                except OSError:
                    pass
    print(f"  Archives directory: {archives_path}")
    print(f"  Archives found: {archives_found}/{len(datasets)}")

    # Generate STAC
    print(f"\nGenerating STAC catalog...")
    n_coll, n_datasets, n_files = generate_stac(datasets, output_path, show_progress=True)
    print(f"\n  Collections: {n_coll}")
    print(f"  Dataset items: {n_datasets}")
    print(f"  File items: {n_files}")

    # =========================================================================
    # DATA QUALITY REPORT
    # =========================================================================
    print(f"\n{'='*60}")
    print("DATA QUALITY REPORT")
    print(f"{'='*60}")

    # 1. Excel entries with no folders/files
    no_folder = [ds for ds in datasets if not ds.folder_path]
    no_files = [ds for ds in datasets if ds.folder_path and len(ds.files) == 0]
    print(f"\n1. EXCEL ENTRIES WITHOUT DATA ({len(no_folder) + len(no_files)} total)")
    print(f"   {'─'*50}")
    if no_folder:
        print(f"   No folder found ({len(no_folder)}):")
        for ds in no_folder[:20]:
            print(f"     - {ds.code}: {ds.sensor} - {ds.dataset}"[:75])
        if len(no_folder) > 20:
            print(f"     ... and {len(no_folder) - 20} more")
    if no_files:
        print(f"   Folder exists but empty ({len(no_files)}):")
        for ds in no_files[:10]:
            print(f"     - {ds.code}: {ds.folder_path}")
        if len(no_files) > 10:
            print(f"     ... and {len(no_files) - 10} more")
    if not no_folder and not no_files:
        print("   ✓ All Excel entries have associated data")

    # 2. Folders that exist but aren't in Excel
    excel_codes = {ds.code for ds in datasets}
    folder_codes = set(code_to_folder.keys())
    orphan_folders = folder_codes - excel_codes
    print(f"\n2. FOLDERS NOT IN EXCEL ({len(orphan_folders)} total)")
    print(f"   {'─'*50}")
    if orphan_folders:
        for code in sorted(orphan_folders)[:20]:
            folder = code_to_folder.get(code, "?")
            print(f"     - {code}: {folder}")
        if len(orphan_folders) > 20:
            print(f"     ... and {len(orphan_folders) - 20} more")
    else:
        print("   ✓ All folders have Excel entries")

    # 3. Datasets with files but no geometry
    has_files_no_geo = [ds for ds in datasets if len(ds.files) > 0 and not ds.bbox]
    print(f"\n3. DATASETS WITH FILES BUT NO GEOMETRY ({len(has_files_no_geo)} total)")
    print(f"   {'─'*50}")
    if has_files_no_geo:
        for ds in has_files_no_geo[:20]:
            file_types = {}
            for fi in ds.files:
                ext = fi.path.suffix.lower()
                file_types[ext] = file_types.get(ext, 0) + 1
            types_str = ", ".join(f"{v}x {k}" for k, v in file_types.items())
            print(f"     - {ds.code}: {len(ds.files)} files ({types_str})")
        if len(has_files_no_geo) > 20:
            print(f"     ... and {len(has_files_no_geo) - 20} more")
    else:
        print("   ✓ All datasets with files have geometry")

    # 4. Datasets with files but no .zip archive
    archives_path = data_path / "archives"
    if not archives_path.exists():
        # Try alternative locations
        for alt in [data_path.parent / "archives", output_path.parent / "archives"]:
            if alt.exists():
                archives_path = alt
                break

    has_files_no_zip = []
    has_zip = []
    for ds in datasets:
        if len(ds.files) > 0:
            zip_path = archives_path / f"{ds.code}.zip"
            if zip_path.exists():
                # Store actual archive size on dataset
                try:
                    ds.archive_size = zip_path.stat().st_size
                except OSError:
                    ds.archive_size = 0
                has_zip.append((ds, zip_path))
            else:
                has_files_no_zip.append(ds)

    print(f"\n4. DATASETS WITH FILES BUT NO .ZIP ARCHIVE ({len(has_files_no_zip)} total)")
    print(f"   {'─'*50}")
    print(f"   Archives directory: {archives_path}")
    print(f"   Archives found: {len(has_zip)}")
    if has_files_no_zip:
        print(f"   Missing archives:")
        for ds in has_files_no_zip[:20]:
            total_size = sum(fi.size for fi in ds.files)
            print(f"     - {ds.code}: {len(ds.files)} files, {total_size / (1024*1024):.1f} MB (uncompressed)")
        if len(has_files_no_zip) > 20:
            print(f"     ... and {len(has_files_no_zip) - 20} more")
    else:
        print("   ✓ All datasets with files have .zip archives")

    # 5. Archive size accuracy
    print(f"\n5. ARCHIVE SIZE ACCURACY")
    print(f"   {'─'*50}")
    if has_zip:
        size_mismatches = []
        for ds, zip_path in has_zip[:50]:  # Check first 50
            zip_size = zip_path.stat().st_size
            uncompressed = sum(fi.size for fi in ds.files)
            # Warn if zip is larger than uncompressed (shouldn't happen) or < 10% (unusual compression)
            if uncompressed > 0:
                ratio = zip_size / uncompressed
                if ratio > 1.1:
                    size_mismatches.append((ds.code, zip_size, uncompressed, "zip larger than files"))
        if size_mismatches:
            print(f"   Unusual archive sizes ({len(size_mismatches)}):")
            for code, zs, us, reason in size_mismatches[:10]:
                print(f"     - {code}: zip={zs/(1024*1024):.1f}MB, files={us/(1024*1024):.1f}MB ({reason})")
        else:
            print("   ✓ Archive sizes look reasonable")
    else:
        print("   (No archives to check)")

    # Summary
    print(f"\n{'─'*60}")
    print("SUMMARY")
    print(f"{'─'*60}")
    total = len(datasets)
    with_folders = len(datasets) - len(no_folder)
    with_files = len([ds for ds in datasets if len(ds.files) > 0])
    with_archives = len(has_zip)
    total_files = sum(len(ds.files) for ds in datasets)
    total_size = sum(fi.size for ds in datasets for fi in ds.files)
    files_with_geo = sum(1 for ds in datasets for fi in ds.files if fi.geometry)

    print(f"  Excel entries:           {total}")
    print(f"  With data folders:       {with_folders:4d}  ({100*with_folders/total:.0f}%)")
    print(f"  With files:              {with_files:4d}  ({100*with_files/total:.0f}%)")
    print(f"  With geometry:           {datasets_with_geo:4d}  ({100*datasets_with_geo/total:.0f}%)")
    print(f"  With .zip archives:      {with_archives:4d}  ({100*with_archives/total:.0f}%)")
    print(f"  Orphan folders:          {len(orphan_folders):4d}")
    print(f"  Total files:             {total_files:4d}")
    print(f"  Files with geometry:     {files_with_geo:4d}  ({100*files_with_geo/max(total_files,1):.0f}%)")
    print(f"  Total size:              {total_size / (1024*1024*1024):.2f} GB")
    print(f"{'='*60}\n")

    # Comprehensive Validation
    print(f"\nValidating STAC catalog...")
    errors = []
    warnings = []

    # Expected bounds for Blatten area (WGS84)
    EXPECTED_BOUNDS = {
        "min_lon": 7.5, "max_lon": 8.5,
        "min_lat": 46.0, "max_lat": 47.0
    }

    def validate_bbox(bbox, item_id):
        """Validate a bounding box."""
        issues = []
        if not bbox or len(bbox) != 4:
            issues.append(f"{item_id}: invalid bbox format")
            return issues

        min_lon, min_lat, max_lon, max_lat = bbox

        # Check order (min < max)
        if min_lon >= max_lon:
            issues.append(f"{item_id}: bbox min_lon ({min_lon}) >= max_lon ({max_lon})")
        if min_lat >= max_lat:
            issues.append(f"{item_id}: bbox min_lat ({min_lat}) >= max_lat ({max_lat})")

        # Check within expected region
        if not (EXPECTED_BOUNDS["min_lon"] <= min_lon <= EXPECTED_BOUNDS["max_lon"]):
            issues.append(f"{item_id}: bbox min_lon ({min_lon}) outside expected range")
        if not (EXPECTED_BOUNDS["min_lat"] <= min_lat <= EXPECTED_BOUNDS["max_lat"]):
            issues.append(f"{item_id}: bbox min_lat ({min_lat}) outside expected range")
        if not (EXPECTED_BOUNDS["min_lon"] <= max_lon <= EXPECTED_BOUNDS["max_lon"]):
            issues.append(f"{item_id}: bbox max_lon ({max_lon}) outside expected range")
        if not (EXPECTED_BOUNDS["min_lat"] <= max_lat <= EXPECTED_BOUNDS["max_lat"]):
            issues.append(f"{item_id}: bbox max_lat ({max_lat}) outside expected range")

        # Check reasonable size (not too small, not too large)
        width = max_lon - min_lon
        height = max_lat - min_lat
        if width > 0.5 or height > 0.5:
            issues.append(f"{item_id}: bbox suspiciously large ({width:.4f} x {height:.4f} degrees)")

        return issues

    def validate_geometry(geom, item_id):
        """Validate GeoJSON geometry."""
        issues = []
        if not geom:
            return issues  # null geometry is allowed

        if "type" not in geom:
            issues.append(f"{item_id}: geometry missing 'type'")
            return issues

        if geom["type"] not in ["Point", "Polygon", "MultiPolygon", "LineString"]:
            issues.append(f"{item_id}: unknown geometry type '{geom['type']}'")

        if "coordinates" not in geom:
            issues.append(f"{item_id}: geometry missing 'coordinates'")

        return issues

    # 1. Validate catalog.json
    print("  Checking catalog.json...")
    catalog_file = output_path / "catalog.json"
    if not catalog_file.exists():
        errors.append("catalog.json does not exist")
    else:
        try:
            with open(catalog_file) as f:
                catalog = json.load(f)
            for field in ["type", "id", "stac_version", "description", "links"]:
                if field not in catalog:
                    errors.append(f"catalog.json missing required field '{field}'")
            if catalog.get("type") != "Catalog":
                errors.append(f"catalog.json type should be 'Catalog', got '{catalog.get('type')}'")
            if catalog.get("stac_version") != STAC_VERSION:
                warnings.append(f"catalog.json stac_version is '{catalog.get('stac_version')}', expected '{STAC_VERSION}'")
        except json.JSONDecodeError as e:
            errors.append(f"catalog.json invalid JSON: {e}")

    # 2. Validate collections.json
    print("  Checking collections.json...")
    collections_file = output_path / "collections.json"
    if not collections_file.exists():
        errors.append("collections.json does not exist")
    else:
        try:
            with open(collections_file) as f:
                coll_data = json.load(f)
            collections_list = coll_data.get("collections", [])
            print(f"    Found {len(collections_list)} collections")

            for coll in collections_list:
                cid = coll.get("id", "unknown")
                for field in ["type", "id", "stac_version", "description", "license", "extent"]:
                    if field not in coll:
                        errors.append(f"Collection '{cid}' missing required field '{field}'")

                # Validate extent
                extent = coll.get("extent", {})
                spatial = extent.get("spatial", {})
                temporal = extent.get("temporal", {})

                if not spatial.get("bbox"):
                    errors.append(f"Collection '{cid}' missing spatial.bbox")
                else:
                    for bbox in spatial["bbox"]:
                        bbox_errors = validate_bbox(bbox, f"Collection '{cid}'")
                        errors.extend(bbox_errors)

                if not temporal.get("interval"):
                    warnings.append(f"Collection '{cid}' missing temporal.interval")

        except json.JSONDecodeError as e:
            errors.append(f"collections.json invalid JSON: {e}")

    # 3. Validate all items
    print("  Checking all_items.json...")
    items_file = output_path / "all_items.json"
    if not items_file.exists():
        errors.append("all_items.json does not exist")
    else:
        try:
            with open(items_file) as f:
                items_data = json.load(f)

            features = items_data.get("features", [])
            print(f"    Found {len(features)} items")

            # Counters
            null_geometry = 0
            null_bbox = 0
            missing_datetime = 0
            missing_collection = 0
            missing_assets = 0
            dataset_items = 0
            file_items = 0

            for item in features:
                item_id = item.get("id", "unknown")

                # Check required STAC fields
                for field in ["type", "stac_version", "id", "properties", "links", "assets"]:
                    if field not in item:
                        errors.append(f"Item '{item_id}' missing required field '{field}'")

                if item.get("type") != "Feature":
                    errors.append(f"Item '{item_id}' type should be 'Feature'")

                # Check geometry
                geom = item.get("geometry")
                if geom is None:
                    null_geometry += 1
                else:
                    geom_errors = validate_geometry(geom, item_id)
                    errors.extend(geom_errors)

                # Check bbox
                bbox = item.get("bbox")
                if bbox is None:
                    null_bbox += 1
                else:
                    bbox_errors = validate_bbox(bbox, item_id)
                    errors.extend(bbox_errors)

                # Check properties
                props = item.get("properties", {})
                if props.get("datetime") is None and props.get("start_datetime") is None:
                    missing_datetime += 1

                # Check collection
                if not item.get("collection"):
                    missing_collection += 1

                # Check assets
                assets = item.get("assets", {})
                if not assets:
                    missing_assets += 1

                # Count item types
                if "-" in item_id and item_id.split("-")[-1].isdigit():
                    file_items += 1
                else:
                    dataset_items += 1

            # Summary
            print(f"    Dataset items: {dataset_items}")
            print(f"    File items: {file_items}")
            files_with_geo = file_items - sum(1 for f in features if "-" in f.get("id", "") and f.get("geometry") is None)
            print(f"    File items with geometry: {files_with_geo} ({100*files_with_geo//max(file_items,1)}%)")

            if null_geometry > 0:
                print(f"    Items with null geometry: {null_geometry}")
            if null_bbox > 0:
                print(f"    Items with null bbox: {null_bbox}")
            if missing_datetime > 0:
                warnings.append(f"{missing_datetime} items missing datetime")
            if missing_collection > 0:
                errors.append(f"{missing_collection} items missing collection reference")
            if missing_assets > 0:
                warnings.append(f"{missing_assets} items have no assets")

        except json.JSONDecodeError as e:
            errors.append(f"all_items.json invalid JSON: {e}")

    # 4. Check collection item files exist
    print("  Checking collection item files...")
    collections_dir = output_path / "collections"
    if collections_dir.exists():
        coll_files = list(collections_dir.glob("*.json"))
        items_files = [f for f in coll_files if f.name.endswith("_items.json")]
        coll_only = [f for f in coll_files if not f.name.endswith("_items.json")]
        print(f"    Collection files: {len(coll_only)}")
        print(f"    Collection item files: {len(items_files)}")

        # Verify each collection has items file
        for cf in coll_only:
            cid = cf.stem
            items_f = collections_dir / f"{cid}_items.json"
            if not items_f.exists():
                warnings.append(f"Collection '{cid}' missing items file")

    # 5. Cross-reference check
    print("  Cross-referencing...")
    if 'catalog' in dir() and 'collections_list' in dir():
        catalog_links = {l.get("href", "").split("/")[-1] for l in catalog.get("links", []) if l.get("rel") == "child"}
        collection_ids = {c.get("id") for c in collections_list}

        missing_in_catalog = collection_ids - catalog_links
        if missing_in_catalog:
            errors.append(f"Collections not linked in catalog: {missing_in_catalog}")

    # 6. S3 validation (optional)
    if args.validate_s3:
        print(f"  Validating files at {args.validate_s3}...")
        try:
            import urllib.request
            import urllib.error

            # Get all asset hrefs from items (skip archives)
            asset_paths = []
            if 'features' in dir():
                for item in features:
                    for asset_key, asset in item.get("assets", {}).items():
                        if asset_key == "archive":
                            continue  # Skip archive assets
                        href = asset.get("href", "")
                        if href.startswith("${S3_BASE_URL}/"):
                            path = href.replace("${S3_BASE_URL}/", "")
                            asset_paths.append(path)

            print(f"    Checking {len(asset_paths)} file assets...")

            # Check files exist using HEAD requests (parallel)
            base_url = args.validate_s3.rstrip("/")
            missing_files = []
            found_files = 0

            def check_file(path):
                url = f"{base_url}/{path}"
                try:
                    req = urllib.request.Request(url, method='HEAD')
                    with urllib.request.urlopen(req, timeout=5) as resp:
                        return (path, resp.status == 200)
                except urllib.error.HTTPError as e:
                    return (path, False)
                except Exception:
                    return (path, False)

            with ThreadPoolExecutor(max_workers=args.workers) as executor:
                futures = {executor.submit(check_file, path): path for path in asset_paths}
                with progress_bar(total=len(futures), desc="Checking S3 files") as pbar:
                    for future in as_completed(futures):
                        path, exists = future.result()
                        if exists:
                            found_files += 1
                        else:
                            missing_files.append(path)
                        pbar.update(1)

            print(f"\n    Files found: {found_files}/{len(asset_paths)}")

            if missing_files:
                print(f"    Missing files: {len(missing_files)}")
                for path in missing_files[:10]:
                    warnings.append(f"File not found: {path}")
                if len(missing_files) > 10:
                    warnings.append(f"... and {len(missing_files) - 10} more files not found on S3")
            else:
                print(f"    All catalog files exist ✓")

        except Exception as e:
            warnings.append(f"S3 validation failed: {e}")

    # Report results
    print(f"\n  Validation Results:")
    print(f"  {'─' * 40}")

    if errors:
        print(f"  ERRORS ({len(errors)}):")
        for err in errors[:20]:  # Show first 20
            print(f"    ✗ {err}")
        if len(errors) > 20:
            print(f"    ... and {len(errors) - 20} more errors")

    if warnings:
        print(f"  WARNINGS ({len(warnings)}):")
        for warn in warnings[:10]:  # Show first 10
            print(f"    ⚠ {warn}")
        if len(warnings) > 10:
            print(f"    ... and {len(warnings) - 10} more warnings")

    if not errors and not warnings:
        print(f"    ✓ All validations passed!")
    elif not errors:
        print(f"    ✓ No errors (but {len(warnings)} warnings)")

    print(f"\nDone. Output: {output_path}")

    # Exit with error code if there are errors
    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
